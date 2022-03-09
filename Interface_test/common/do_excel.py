from openpyxl import load_workbook
from common.project_path import conf_path, test_data_path
from common.read_config import ReadConfig


class DoExcel:
    """该类完成测试数据的读取以及测试结果的写回"""

    def __init__(self, file_name, sheet_name):
        self.test_data = []
        self.file_name = file_name  # Excel工作簿文件名或地址
        self.sheet_name = sheet_name  # 表单名

    def read_data(self, section):  # section 配置文件里面的片段名 可以根据你的指定来执行具体的用例
        """从Excel读取数据，返回json数据"""
        # 从配置文件里面控制读取哪些用例
        case_id = ReadConfig(conf_path).get_data(section, 'case_id')
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]

        for i in range(2, sheet.max_row + 1):
            row_data = dict()
            row_data['CaseId'] = sheet.cell(i, 1).value
            row_data['Module'] = sheet.cell(i, 2).value
            row_data['Title'] = sheet.cell(i, 3).value
            row_data['Headers'] = sheet.cell(i, 4).value
            row_data['Method'] = sheet.cell(i, 5).value
            row_data['Params'] = sheet.cell(i, 6).value
            row_data['sql'] = sheet.cell(i, 7).value
            row_data['ExpectedResult'] = sheet.cell(i, 8).value
            self.test_data.append(row_data)
        wb.close()
        final_data = []  # 空列表 存储最终的测试用例数据
        if case_id == 'all':  # 如果case_id==all 那就获取所有的用例数据
            final_data = self.test_data  # 把测试用例赋值给final_data这个变量
        else:  # 否则 如果是列表 那就获取列表里面指定id的用例的数据
            for i in case_id:  # 遍历case_id 里面的值
                final_data.append(self.test_data[i - 1])
        return final_data

    def write_back(self, row, col, value):
        """写回测试结果到Excel中"""
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        sheet.cell(row, col).value = value
        wb.save(self.file_name)
        wb.close()  # 关闭文件的动作


if __name__ == '__main__':
    test_data = DoExcel(test_data_path, 'rst').read_data('CASE')  # 获取测试数据
    print(test_data)
    print('-' * 300)
    for i in test_data:
        print(i)
